import datetime
from io import BytesIO
import openpyxl


TEMPLATE_PATH = "template.xlsx"

WEEK_CELL = "I5"
START_ROW = 10

COL_FNAME = 1
COL_LNAME = 2
COL_CC = 3
COL_PCODE = 4
COL_DAYS_START = 5

DATES_ROW = 8


def generate_excel_report(
    workers_data: dict,
    week_number: int,
    start_of_week: datetime.date
) -> BytesIO:

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # Номер недели
    ws[WEEK_CELL] = week_number


    # Заполняем даты
    for day_idx in range(7):

        day_date = start_of_week + datetime.timedelta(days=day_idx)

        cell = ws.cell(
            row=DATES_ROW,
            column=COL_DAYS_START + day_idx
        )

        cell.value = day_date
        cell.number_format = "DD-MM-YYYY"


    current_row = START_ROW


    # Сортировка работников
    sorted_workers = sorted(
        workers_data.items(),
        key=lambda x: x[0].full_name
    )


    for worker, dates in sorted_workers:

        name_parts = worker.full_name.strip().split(" ", 1)

        first_name = name_parts[0]
        last_name = (
            name_parts[1]
            if len(name_parts) > 1
            else ""
        )


        ws.cell(
            row=current_row,
            column=COL_FNAME,
            value=first_name
        )

        ws.cell(
            row=current_row,
            column=COL_LNAME,
            value=last_name
        )

        ws.cell(
            row=current_row,
            column=COL_CC,
            value=worker.cc or ""
        )

        ws.cell(
            row=current_row,
            column=COL_PCODE,
            value=worker.projectcode
        )


        # Часы по дням
        for day_idx in range(7):

            day_date = (
                start_of_week
                + datetime.timedelta(days=day_idx)
            )

            hours = dates.get(day_date, 0)


            ws.cell(
                row=current_row,
                column=COL_DAYS_START + day_idx,
                value=hours
            )


        current_row += 1


    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return output